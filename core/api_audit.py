import json
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import AuditLog, User

PAGE_SIZE = 50

FIELD_LABELS = {
    'last_name': 'Фамилия', 'first_name': 'Имя', 'middle_name': 'Отчество',
    'full_name': 'ФИО',
    'phone': 'Телефон', 'email': 'Email',
    'date_of_birth': 'Дата рождения', 'birth_date': 'Дата рождения',
    'address': 'Адрес', 'status': 'Статус',
    'group': 'Группа', 'group_id': 'Группа', 'group_number': 'Номер группы',
    'name': 'Название', 'full_name_ru': 'Полное название', 'short_name': 'Аббревиатура',
    'code': 'Код', 'description': 'Описание',
    'year': 'Год начала',
    'position': 'Должность', 'position_id': 'Должность',
    'faculty': 'Факультет', 'faculty_id': 'Факультет',
    'curator': 'Куратор', 'curator_id': 'Куратор',
    'headteacher': 'Классный руководитель', 'headteacher_id': 'Классный руководитель',
    'employee': 'Сотрудник', 'employee_id': 'Сотрудник',
    'student': 'Студент', 'student_id': 'Студент',
    'parent': 'Опекун', 'parent_id': 'Опекун',
    'subject': 'Предмет', 'subject_id': 'Предмет',
    'relation_type': 'Тип связи',
    'username': 'Логин',
    'role': 'Роль', 'is_active': 'Активен',
    'repr': 'Объект', 'reason': 'Причина',
    'institution': 'Организация', 'institution_id': 'Организация',
    'founded_date': 'Дата основания',
}

ACTION_MAP = {
    'created': ('create', 'Создал', 'badge-ok'),
    'updated': ('update', 'Изменил', 'badge-warn'),
    'deleted': ('delete', 'Удалил', 'badge-bad'),
    'transferred': ('transfer', 'Перевёл студента', 'badge-info'),
}

ROLE_LABELS = {
    'owner': 'Суперадмин',
    'admin': 'Администратор',
    'teacher': 'Преподаватель',
}

OBJ_TYPE_LABELS = {
    'Student': 'Студент', 'Employee': 'Сотрудник', 'Group': 'Группа',
    'Faculty': 'Факультет', 'Parent': 'Опекун', 'User': 'Пользователь',
    'StudentParent': 'Связь опекун-студент', 'GroupSubjectEmployee': 'Назначение предмета',
}


def _extract_name(new_data_str, old_data_str):
    for data_str in (new_data_str, old_data_str):
        if not data_str:
            continue
        try:
            d = json.loads(data_str)
            for key in ('full_name', 'name', 'username'):
                if d.get(key):
                    return d[key]
        except Exception:
            pass
    return None


def _compute_changes(old_data_str, new_data_str):
    try:
        old = json.loads(old_data_str) if old_data_str else {}
    except Exception:
        old = {}
    try:
        new = json.loads(new_data_str) if new_data_str else {}
    except Exception:
        new = {}
    changes = []
    for key in set(list(old.keys()) + list(new.keys())):
        old_val = old.get(key)
        new_val = new.get(key)
        if old_val != new_val:
            changes.append({
                'key': key,
                'label': FIELD_LABELS.get(key, key),
                'from': old_val if old_val is not None else '-',
                'to': new_val if new_val is not None else '-',
            })
    return changes


def _serialize(log):
    action_key, action_label, action_cls = ACTION_MAP.get(log.action, ('update', log.action, 'badge-warn'))
    u = log.user
    user_login = (u.username if u else None) or '-'
    user_role = ROLE_LABELS.get(u.role, u.role or '-') if u else '-'
    user_position = None
    if u and u.employee_id:
        try:
            pos = u.employee.position
            user_position = pos.name if pos else None
        except Exception:
            pass

    obj_name = _extract_name(log.new_data, log.old_data) or ''
    obj_type = OBJ_TYPE_LABELS.get(log.object_type, log.object_type)

    return {
        'id': log.pk,
        'ts': log.created_at.strftime('%d.%m.%Y %H:%M:%S'),
        'user': user_login,
        'userName': (u.display_name or u.username if u else None) or '-',
        'role': user_role,
        'userPosition': user_position,
        'action': action_key,
        'label': action_label,
        'cls': action_cls,
        'obj': obj_name or obj_type,
        'obj_name': obj_name,
        'obj_type': obj_type,
        'object_id': log.object_id,
        'object_type_raw': log.object_type,
        'changes': _compute_changes(log.old_data, log.new_data),
    }


def _apply_filters(qs, request):
    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(object_type__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__display_name__icontains=search)
        )

    action = request.GET.get('action', '')
    if action:
        db_action = {'create': 'created', 'update': 'updated', 'delete': 'deleted', 'transfer': 'transferred'}.get(action)
        if db_action:
            qs = qs.filter(action=db_action)

    period = request.GET.get('period', 'all')
    now = timezone.now()
    if period == 'day':
        qs = qs.filter(created_at__gte=now - timedelta(hours=24))
    elif period == 'week':
        qs = qs.filter(created_at__gte=now - timedelta(days=7))
    elif period == 'month':
        qs = qs.filter(created_at__gte=now - timedelta(days=30))

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    user_id = request.GET.get('user_id', '')
    if user_id:
        try:
            qs = qs.filter(user_id=int(user_id))
        except (ValueError, TypeError):
            pass

    role = request.GET.get('role', '')
    if role:
        qs = qs.filter(user__role=role)

    object_type = request.GET.get('object_type', '')
    if object_type:
        qs = qs.filter(object_type=object_type)

    object_id = request.GET.get('object_id', '')
    if object_id:
        try:
            qs = qs.filter(object_id=int(object_id))
        except (ValueError, TypeError):
            pass

    return qs


class AuditLogView(APIView):
    def get(self, request):
        if request.user.role not in ('owner', 'admin'):
            return Response({'error': 'Доступ запрещён'}, status=403)
        institution = request.user.institution
        if not institution:
            return Response({'results': [], 'count': 0, 'num_pages': 1})

        qs = AuditLog.objects.filter(institution=institution).select_related(
            'user', 'user__employee', 'user__employee__position'
        )
        qs = _apply_filters(qs, request)

        count = qs.count()
        try:
            page = max(1, int(request.GET.get('page', 1)))
        except (ValueError, TypeError):
            page = 1
        num_pages = max(1, (count + PAGE_SIZE - 1) // PAGE_SIZE)
        page = min(page, num_pages)
        offset = (page - 1) * PAGE_SIZE

        logs = qs.order_by('-created_at')[offset:offset + PAGE_SIZE]
        return Response({
            'results': [_serialize(log) for log in logs],
            'count': count,
            'num_pages': num_pages,
        })


class AuditLogExportView(APIView):
    def get(self, request):
        if request.user.role not in ('owner', 'admin'):
            return Response({'error': 'Доступ запрещён'}, status=403)
        institution = request.user.institution
        if not institution:
            return Response({'error': 'Нет организации'}, status=400)

        qs = AuditLog.objects.filter(institution=institution).select_related(
            'user', 'user__employee', 'user__employee__position'
        )
        qs = _apply_filters(qs, request)

        sort_map = {
            'ts': 'created_at',
            'user': 'user__username',
            'label': 'action',
            'obj': 'object_type',
        }
        sort_key = request.GET.get('sort_key', 'ts')
        sort_dir = request.GET.get('sort_dir', 'desc')
        order_field = sort_map.get(sort_key, 'created_at')
        prefix = '' if sort_dir == 'asc' else '-'
        qs = qs.order_by(f'{prefix}{order_field}')

        rows = [_serialize(log) for log in qs[:10000]]

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Журнал изменений'

        headers = ['N', 'Дата и время', 'Логин', 'ФИО', 'Роль', 'Должность', 'Действие', 'Тип объекта', 'Объект', 'Поле', 'Было', 'Стало']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9D9D9')

        for i, a in enumerate(rows, 1):
            base = [
                i,
                a['ts'],
                a['user'],
                a['userName'],
                a['role'],
                a.get('userPosition') or '',
                a['label'],
                a['obj_type'],
                a['obj_name'] or a['obj'],
            ]
            changes = a.get('changes', [])
            if changes:
                for ch in changes:
                    ws.append(base + [ch['label'], str(ch['from']), str(ch['to'])])
            else:
                ws.append(base + ['', '', ''])

        col_widths = [5, 20, 16, 24, 16, 20, 14, 18, 30, 22, 22, 22]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
        wb.save(response)
        return response


class AuditRollbackView(APIView):
    _ROLLBACK_FIELDS = {
        'Student':  ['last_name', 'first_name', 'middle_name', 'phone', 'email', 'birth_date', 'status'],
        'Employee': ['last_name', 'first_name', 'middle_name', 'phone', 'email', 'birth_date'],
        'Faculty':  ['full_name', 'short_name'],
        'Group':    ['year'],
        'Parent':   ['last_name', 'first_name', 'middle_name', 'phone', 'email', 'birth_date'],
    }

    def _get_obj(self, object_type, object_id, institution):
        from .models import Student, Employee, Faculty, Group, Parent
        try:
            if object_type == 'Student':
                return Student.objects.get(pk=object_id, faculty__institution=institution)
            if object_type == 'Employee':
                return Employee.objects.get(pk=object_id, institution=institution)
            if object_type == 'Faculty':
                return Faculty.objects.get(pk=object_id, institution=institution)
            if object_type == 'Group':
                return Group.objects.get(pk=object_id, faculty__institution=institution)
            if object_type == 'Parent':
                return Parent.objects.get(pk=object_id, institution=institution)
        except Exception:
            pass
        return None

    def post(self, request, pk):
        if request.user.role not in ('owner', 'admin'):
            return Response({'error': 'Доступ запрещён'}, status=403)
        institution = request.user.institution
        try:
            log = AuditLog.objects.get(pk=pk, institution=institution)
        except AuditLog.DoesNotExist:
            return Response({'error': 'Запись не найдена'}, status=404)

        if log.action != 'updated':
            return Response({'error': 'Откат доступен только для записей об изменении'}, status=400)

        try:
            old_data = json.loads(log.old_data) if log.old_data else {}
        except Exception:
            old_data = {}

        if not old_data:
            return Response({'error': 'Нет данных для отката'}, status=400)

        obj = self._get_obj(log.object_type, log.object_id, institution)
        if obj is None:
            return Response({'error': 'Объект не найден - возможно, он был удалён'}, status=404)

        allowed = self._ROLLBACK_FIELDS.get(log.object_type, [])
        if not allowed:
            return Response({'error': 'Откат для этого типа объектов не поддерживается'}, status=400)

        before_rollback = {}
        applied = False
        for field in allowed:
            if field not in old_data:
                continue
            val = old_data[field]
            if val in ('None', '-'):
                val = None
            if field == 'year' and val is not None:
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    continue
            current = getattr(obj, field, None)
            before_rollback[field] = str(current) if current is not None else None
            current_str = str(current) if current is not None else None
            old_str = str(val) if val is not None else None
            if current_str != old_str:
                setattr(obj, field, val)
                applied = True

        if not applied:
            return Response({'error': 'Нечего откатывать - данные уже совпадают'}, status=400)

        obj.save()
        after_rollback = {f: str(getattr(obj, f)) if getattr(obj, f) is not None else None
                         for f in before_rollback}

        from .utils import log_action
        log_action(request.user, 'updated', obj,
                   old_data=before_rollback,
                   new_data=after_rollback,
                   institution=institution)

        return Response({'ok': True})


class AuditLogUsersView(APIView):
    def get(self, request):
        if request.user.role not in ('owner', 'admin'):
            return Response([])
        institution = request.user.institution
        if not institution:
            return Response([])
        user_ids = (
            AuditLog.objects.filter(institution=institution, user__isnull=False)
            .values_list('user_id', flat=True).distinct()
        )
        users = User.objects.filter(pk__in=user_ids)
        return Response([
            {
                'value': u.pk,
                'label': u.username,
                'sub': ROLE_LABELS.get(u.role, u.role or ''),
            }
            for u in users
        ])
